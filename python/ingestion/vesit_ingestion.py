import os
import sys
import datetime
import openpyxl
import re

# Add backend directory to sys.path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../../backend')))
from app.db.session import SessionLocal, engine
from app.db.base import Base
from app.models.schema import (
    DataSource, Institute, Campus, EnergyConsumption, Appliance,
    EmissionFactor, CarbonEmission, CarbonCalculationRun, CarbonCalculationLog
)

MONTH_NAMES = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December"
}

def seed_vesit_entities(db):
    """Ensure VESIT Institute and Campus exist in database."""
    institute = db.query(Institute).filter(Institute.institute_name.ilike("%VESIT%") | Institute.institute_name.ilike("%V.E.S%")).first()
    if not institute:
        institute = Institute(
            institute_name="V.E.S. Institute of Technology (VESIT)",
            location="Chembur, Mumbai, Maharashtra, India",
            area=36000.0
        )
        db.add(institute)
        db.flush()
        print(f"Created Institute: {institute.institute_name} (ID={institute.institute_id})")

    campus = db.query(Campus).filter(Campus.institute_id == institute.institute_id).first()
    if not campus:
        campus = Campus(
            institute_id=institute.institute_id,
            campus_name="VESIT Chembur Campus",
            location="Collector's Colony, Chembur, Mumbai",
            area=36000.0
        )
        db.add(campus)
        db.flush()
        print(f"Created Campus: {campus.campus_name} (ID={campus.campus_id})")

    # Ensure Grid Emission Factor exists (CEA India default: 0.82 kgCO2e/kWh)
    grid_ef = db.query(EmissionFactor).filter(
        EmissionFactor.category.ilike("%Electricity%") | EmissionFactor.subcategory.ilike("%Grid%")
    ).first()
    if not grid_ef:
        grid_ef = EmissionFactor(
            category="Scope 2",
            subcategory="Grid Electricity",
            activity_unit="kWh",
            factor_value=0.82,
            factor_unit="kgCO2e/kWh",
            scope="Scope 2",
            source_name="Central Electricity Authority (CEA), Ministry of Power, Govt of India",
            source_reference="CO2 Baseline Database for the Indian Power Sector, CEA",
            notes="Standard average grid emission factor for Western Regional Grid / National Grid (India)."
        )
        db.add(grid_ef)
        db.flush()
        print(f"Created Grid Emission Factor: {grid_ef.factor_value} {grid_ef.factor_unit}")

    db.commit()
    return institute, campus, grid_ef

def parse_electricity_sheet(sheet, year_name, db, ds, institute, campus):
    """
    Parse a single yearly sheet from Electricity usage Data 2022 to 2026.xlsx
    Handles variable column layouts, multi-row headers, A Wing, B Wing, and Construction.
    """
    if sheet.max_row is None or sheet.max_row <= 2:
        print(f"Sheet {year_name} is empty. Skipping.")
        return 0

    # Locate the header row containing "A WING" and "B WING"
    wing_header_row_idx = None
    sub_header_row_idx = None

    for r in range(1, min(10, sheet.max_row + 1)):
        row_vals = [str(sheet.cell(r, c).value or "").upper() for c in range(1, min(15, sheet.max_column + 1))]
        if any("A WING" in v or "A-WING" in v for v in row_vals):
            wing_header_row_idx = r
            sub_header_row_idx = r + 1
            break

    if not wing_header_row_idx:
        print(f"Could not locate A/B Wing header in sheet {year_name}. Inspecting alternative rows...")
        return 0

    # Map column indices for A Wing, B Wing, and Construction
    max_col = min(sheet.max_column, 20)
    
    a_wing_cols = {}
    b_wing_cols = {}
    const_cols = {}

    current_wing = None
    for c in range(1, max_col + 1):
        top_val = str(sheet.cell(wing_header_row_idx, c).value or "").upper()
        if "A WING" in top_val:
            current_wing = "A Wing"
        elif "B WING" in top_val:
            current_wing = "B Wing"
        elif "CONTS" in top_val or "CONST" in top_val:
            current_wing = "Construction"

        sub_val = str(sheet.cell(sub_header_row_idx, c).value or "").strip().upper()

        if current_wing == "A Wing":
            if "RATE" in sub_val or "PER UNIT" in sub_val:
                a_wing_cols["rate"] = c
            elif "DEMAND" in sub_val or "BILLING" in sub_val:
                a_wing_cols["demand"] = c
            elif "CONSUMED" in sub_val or "KWH" in sub_val or sub_val == "UNIT" or "UNIT CONSUMED" in sub_val:
                a_wing_cols["units"] = c
            elif "AMOUNT" in sub_val or "PAID" in sub_val or "EXPENSE" in sub_val:
                a_wing_cols["amount"] = c
            elif "POWER FACTOR" in sub_val or "PF" in sub_val:
                a_wing_cols["pf"] = c
        elif current_wing == "B Wing":
            if "RATE" in sub_val or "PER UNIT" in sub_val:
                b_wing_cols["rate"] = c
            elif "DEMAND" in sub_val or "BILLING" in sub_val:
                b_wing_cols["demand"] = c
            elif "CONSUMED" in sub_val or "KWH" in sub_val or sub_val == "UNIT" or "UNIT CONSUMED" in sub_val:
                b_wing_cols["units"] = c
            elif "AMOUNT" in sub_val or "PAID" in sub_val or "EXPENSE" in sub_val:
                b_wing_cols["amount"] = c
            elif "POWER FACTOR" in sub_val or "PF" in sub_val:
                b_wing_cols["pf"] = c
        elif current_wing == "Construction":
            if "CONSUMED" in sub_val or "UNIT" in sub_val:
                const_cols["units"] = c
            elif "AMOUNT" in sub_val or "PAID" in sub_val:
                const_cols["amount"] = c

    print(f"Sheet {year_name} mapping: A_Wing={a_wing_cols}, B_Wing={b_wing_cols}, Construction={const_cols}")

    records = []
    # Parse data rows starting after sub_header_row_idx
    for r in range(sub_header_row_idx + 1, sheet.max_row + 1):
        cell_1_val = sheet.cell(r, 1).value
        if cell_1_val is None:
            continue

        # Check if this is a summary row
        cell_1_str = str(cell_1_val).strip().upper()
        if "TOTAL" in cell_1_str or "AVG" in cell_1_str or "AVERAGE" in cell_1_str or "WING" in cell_1_str:
            continue

        # Parse Date
        rec_date = None
        if isinstance(cell_1_val, (datetime.datetime, datetime.date)):
            rec_date = cell_1_val.date() if isinstance(cell_1_val, datetime.datetime) else cell_1_val
        elif isinstance(cell_1_val, str):
            try:
                dt = datetime.datetime.strptime(cell_1_val.strip(), "%Y-%m-%d")
                rec_date = dt.date()
            except:
                continue

        if not rec_date:
            continue

        # 1. Ingest A Wing Record
        if "units" in a_wing_cols:
            raw_units = sheet.cell(r, a_wing_cols["units"]).value
            if raw_units is not None and str(raw_units).strip() != "":
                try:
                    units_val = float(raw_units)
                    raw_amount = sheet.cell(r, a_wing_cols.get("amount", 0)).value if "amount" in a_wing_cols else None
                    amount_val = float(raw_amount) if raw_amount is not None and str(raw_amount).strip() != "" else None
                    raw_demand = sheet.cell(r, a_wing_cols.get("demand", 0)).value if "demand" in a_wing_cols else None
                    demand_val = float(raw_demand) if raw_demand is not None and str(raw_demand).strip() != "" else None
                    raw_rate = sheet.cell(r, a_wing_cols.get("rate", 0)).value if "rate" in a_wing_cols else None
                    rate_val = float(raw_rate) if raw_rate is not None and str(raw_rate).strip() != "" else (amount_val / units_val if amount_val and units_val > 0 else None)
                    raw_pf = sheet.cell(r, a_wing_cols.get("pf", 0)).value if "pf" in a_wing_cols else None
                    pf_val = float(raw_pf) if raw_pf is not None and str(raw_pf).strip() != "" else None

                    rec = EnergyConsumption(
                        institute_id=institute.institute_id,
                        campus_id=campus.campus_id,
                        wing="A Wing",
                        timestamp=datetime.datetime(rec_date.year, rec_date.month, rec_date.day, 0, 0, tzinfo=datetime.timezone.utc),
                        date=rec_date,
                        year=rec_date.year,
                        month=MONTH_NAMES.get(rec_date.month, str(rec_date.month)),
                        month_number=rec_date.month,
                        billing_demand=demand_val,
                        units_consumed_kwh=units_val,
                        amount_paid=amount_val,
                        rate_per_unit=rate_val,
                        power_factor=pf_val,
                        energy_source="Grid Electricity",
                        consumption_value=units_val,
                        unit="kWh",
                        data_source_type="VESIT_ACTUAL",
                        source_dataset_id=ds.source_dataset_id
                    )
                    records.append(rec)
                except Exception as e:
                    print(f"Error parsing A Wing row {r} in {year_name}: {e}")

        # 2. Ingest B Wing Record
        if "units" in b_wing_cols:
            raw_units = sheet.cell(r, b_wing_cols["units"]).value
            if raw_units is not None and str(raw_units).strip() != "":
                try:
                    units_val = float(raw_units)
                    raw_amount = sheet.cell(r, b_wing_cols.get("amount", 0)).value if "amount" in b_wing_cols else None
                    amount_val = float(raw_amount) if raw_amount is not None and str(raw_amount).strip() != "" else None
                    raw_demand = sheet.cell(r, b_wing_cols.get("demand", 0)).value if "demand" in b_wing_cols else None
                    demand_val = float(raw_demand) if raw_demand is not None and str(raw_demand).strip() != "" else None
                    raw_rate = sheet.cell(r, b_wing_cols.get("rate", 0)).value if "rate" in b_wing_cols else None
                    rate_val = float(raw_rate) if raw_rate is not None and str(raw_rate).strip() != "" else (amount_val / units_val if amount_val and units_val > 0 else None)
                    raw_pf = sheet.cell(r, b_wing_cols.get("pf", 0)).value if "pf" in b_wing_cols else None
                    pf_val = float(raw_pf) if raw_pf is not None and str(raw_pf).strip() != "" else None

                    rec = EnergyConsumption(
                        institute_id=institute.institute_id,
                        campus_id=campus.campus_id,
                        wing="B Wing",
                        timestamp=datetime.datetime(rec_date.year, rec_date.month, rec_date.day, 0, 0, tzinfo=datetime.timezone.utc),
                        date=rec_date,
                        year=rec_date.year,
                        month=MONTH_NAMES.get(rec_date.month, str(rec_date.month)),
                        month_number=rec_date.month,
                        billing_demand=demand_val,
                        units_consumed_kwh=units_val,
                        amount_paid=amount_val,
                        rate_per_unit=rate_val,
                        power_factor=pf_val,
                        energy_source="Grid Electricity",
                        consumption_value=units_val,
                        unit="kWh",
                        data_source_type="VESIT_ACTUAL",
                        source_dataset_id=ds.source_dataset_id
                    )
                    records.append(rec)
                except Exception as e:
                    print(f"Error parsing B Wing row {r} in {year_name}: {e}")

        # 3. Ingest Construction Record (if present)
        if "units" in const_cols:
            raw_units = sheet.cell(r, const_cols["units"]).value
            if raw_units is not None and str(raw_units).strip() != "":
                try:
                    units_val = float(raw_units)
                    raw_amount = sheet.cell(r, const_cols.get("amount", 0)).value if "amount" in const_cols else None
                    amount_val = float(raw_amount) if raw_amount is not None and str(raw_amount).strip() != "" else None

                    rec = EnergyConsumption(
                        institute_id=institute.institute_id,
                        campus_id=campus.campus_id,
                        wing="Construction",
                        timestamp=datetime.datetime(rec_date.year, rec_date.month, rec_date.day, 0, 0, tzinfo=datetime.timezone.utc),
                        date=rec_date,
                        year=rec_date.year,
                        month=MONTH_NAMES.get(rec_date.month, str(rec_date.month)),
                        month_number=rec_date.month,
                        units_consumed_kwh=units_val,
                        amount_paid=amount_val,
                        rate_per_unit=amount_val / units_val if amount_val and units_val > 0 else None,
                        energy_source="Grid Electricity",
                        consumption_value=units_val,
                        unit="kWh",
                        data_source_type="VESIT_ACTUAL",
                        source_dataset_id=ds.source_dataset_id
                    )
                    records.append(rec)
                except Exception as e:
                    print(f"Error parsing Construction row {r} in {year_name}: {e}")

    for rec in records:
        db.add(rec)
    db.commit()
    print(f"Successfully ingested {len(records)} monthly records for sheet {year_name}.")
    return len(records)

def ingest_vesit_electricity(excel_path):
    """Main function to ingest VESIT Electricity usage Excel file."""
    print(f"\n==========================================")
    print(f"INGESTING VESIT ELECTRICITY: {excel_path}")
    print(f"==========================================")
    
    if not os.path.exists(excel_path):
        print(f"ERROR: File not found at {excel_path}")
        return

    db = SessionLocal()
    try:
        institute, campus, grid_ef = seed_vesit_entities(db)

        # Register or retrieve DataSource
        filename = os.path.basename(excel_path)
        ds = db.query(DataSource).filter(DataSource.original_filename == filename).first()
        if not ds:
            ds = DataSource(
                source_name="VESIT Official Electricity Usage (2022-2026)",
                source_type="Electricity Usage Bills",
                original_filename=filename,
                is_test_data=False,
                notes="Actual institute monthly billing data for A Wing (350 KVA contract demand) and B Wing (175 KVA contract demand)."
            )
            db.add(ds)
            db.commit()
            db.refresh(ds)

        # Clear any existing records for this data source to ensure idempotency
        db.query(EnergyConsumption).filter(EnergyConsumption.data_source_type == "VESIT_ACTUAL").delete()
        db.commit()

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        total_ingested = 0
        for sheetname in wb.sheetnames:
            print(f"\n--- Processing Sheet: {sheetname} ---")
            sheet = wb[sheetname]
            count = parse_electricity_sheet(sheet, sheetname, db, ds, institute, campus)
            total_ingested += count

        print(f"\n>> TOTAL VESIT Electricity Records Ingested: {total_ingested}")
    finally:
        db.close()

def classify_equipment(name_raw):
    """Categorize appliance names and extract capacity if present."""
    name = str(name_raw or "").strip()
    name_upper = name.upper()

    category = "Other"
    rated_power_kw = None
    capacity_tr = None
    capacity_hp = None

    # Renewable Energy & Trees
    if "SOLAR" in name_upper or "PV" in name_upper:
        category = "Renewable Energy"
    elif "TREE" in name_upper or "PLANT" in name_upper:
        category = "Campus Greenery"

    # Water Systems
    elif "PUMP" in name_upper:
        category = "Water Systems"
    elif "WATER COOLER" in name_upper:
        category = "Water Systems"

    # Vertical Transport
    elif "LIFT" in name_upper or "ELEVATOR" in name_upper:
        category = "Vertical Transport"

    # IT Equipment
    elif name_upper.startswith("PC") or "PCS" in name_upper:
        category = "IT Equipment"
    elif "PRINTER" in name_upper:
        category = "IT Equipment"

    # Ventilation
    elif "FAN" in name_upper:
        category = "Ventilation"

    # Lighting
    elif "TUBE LIGHT" in name_upper or "36 W" in name_upper:
        category = "Lighting"
        rated_power_kw = 0.036
    elif "CFL" in name_upper or "C F L" in name_upper:
        category = "Lighting"
    elif "L E D- 1 W" in name_upper or "LED- 1 W" in name_upper or "LED 1 W" in name_upper:
        category = "Lighting"
        rated_power_kw = 0.001
    elif "15 W" in name_upper:
        category = "Lighting"
        rated_power_kw = 0.015
    elif "20 W" in name_upper:
        category = "Lighting"
        rated_power_kw = 0.020
    elif "40 W" in name_upper:
        category = "Lighting"
        rated_power_kw = 0.040
    elif "LED" in name_upper:
        category = "Lighting"

    # Cooling (ACs)
    elif re.search(r'\b(AC|ACS|A C|A CS|AIR CONDITIONER|CASSETTE|CASSEETTE|TOWER AC)\b', name_upper) or " TR" in name_upper or "TR CAPACITY" in name_upper or "CASSEETTE" in name_upper:
        category = "Cooling"
        if "0.75" in name:
            capacity_tr = 0.75
        elif "1.5" in name:
            capacity_tr = 1.5
        elif "1 TR" in name_upper or " 1 " in name:
            capacity_tr = 1.0
        elif "2 TR" in name_upper or "2TR" in name_upper or " 2 " in name:
            capacity_tr = 2.0
        elif "3 TR" in name_upper or "3TR" in name_upper or " 3 " in name:
            capacity_tr = 3.0

    return category, rated_power_kw, capacity_tr, capacity_hp

def ingest_vesit_appliances(excel_path):
    """Main function to ingest VESIT Appliance Inventory Excel file."""
    print(f"\n==========================================")
    print(f"INGESTING VESIT APPLIANCE INVENTORY: {excel_path}")
    print(f"==========================================")

    if not os.path.exists(excel_path):
        print(f"ERROR: File not found at {excel_path}")
        return

    db = SessionLocal()
    try:
        institute, campus, _ = seed_vesit_entities(db)

        filename = os.path.basename(excel_path)
        ds = db.query(DataSource).filter(DataSource.original_filename == filename).first()
        if not ds:
            ds = DataSource(
                source_name="VESIT Electrical Load & Appliance Inventory",
                source_type="Appliance Inventory",
                original_filename=filename,
                is_test_data=False,
                notes="Official campus electrical load asset inventory for VESIT."
            )
            db.add(ds)
            db.commit()
            db.refresh(ds)

        # Clear existing VESIT appliance inventory records
        db.query(Appliance).filter(Appliance.institute_id == institute.institute_id).delete()
        db.commit()

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

        records = []
        for r in range(4, sheet.max_row + 1):
            eq_name = sheet.cell(r, 2).value
            qty_val = sheet.cell(r, 3).value

            if eq_name is None:
                continue

            eq_name_str = str(eq_name).strip()
            if not eq_name_str:
                continue

            parsed_qty = None
            if qty_val is not None:
                try:
                    parsed_qty = int(float(qty_val))
                except:
                    parsed_qty = None

            category, rated_power_kw, capacity_tr, capacity_hp = classify_equipment(eq_name_str)

            # Operating hours default estimates by category
            hours = 8.0
            if category == "Cooling":
                hours = 7.0
            elif category == "Lighting":
                hours = 9.0
            elif category == "IT Equipment":
                hours = 8.0
            elif category == "Ventilation":
                hours = 8.0
            elif category == "Vertical Transport":
                hours = 12.0
            elif category == "Water Systems":
                hours = 4.0

            app = Appliance(
                institute_id=institute.institute_id,
                appliance_name=eq_name_str,
                appliance_category=category,
                appliance_type=category,
                quantity=parsed_qty if parsed_qty is not None else 0,
                rated_power_kw=rated_power_kw,
                capacity_tr=capacity_tr,
                capacity_hp=capacity_hp,
                operating_hours_per_day=hours,
                source_type="VESIT_ACTUAL",
                source_dataset_id=ds.source_dataset_id
            )
            records.append(app)

        for a in records:
            db.add(a)
        db.commit()
        print(f">> Successfully Ingested {len(records)} Appliance Inventory Records for VESIT.")
    finally:
        db.close()

if __name__ == "__main__":
    elec_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../database/data/vesit/Electricity usage Data 2022 to 2026.xlsx'))
    app_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../database/data/vesit/No. of Appliences.xlsx'))

    ingest_vesit_electricity(elec_path)
    ingest_vesit_appliances(app_path)
